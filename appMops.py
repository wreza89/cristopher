# Bibliotecas
import streamlit as st
import openpyxl

################################################################################################################
def mopMantenimientos():

    workbook = openpyxl.load_workbook('MOP-CRQ.xlsx')
    sheets = workbook.sheetnames
    sheet = workbook.active

    nombreCRQ = st.text_input('Nombre del Mantenimiento:')
    sheet['B9'] = nombreCRQ
    sheet['C11'] = nombreCRQ
    CRQ = nombreCRQ[4:10]
    numCRQ = st.text_input('Numero de CRQ:', value=CRQ)
    if numCRQ.isdigit() == True:
        sheet['D10'] = numCRQ
    else:
        st.warning('Ingrese un numero valido')
    
    impacto = st.text_input('Impacto / Alcance', value='Ninguno')
    sheet['C13'] = impacto

    n2Responsable = st.selectbox('N2', ['Carlos Castillo','Cristhian Vallejo', 'Elizabeth Peñaloza', 'Fabricio Oleas','Luis Chumo', 'Washington Chuqui', 'Patricio Jordán', 'Mauro Caluña', 'Xavier Ortiz'])
    sheet['E18'] = n2Responsable
    if n2Responsable == 'Carlos Castillo':
        sheet['H18'] = 'carlos.castillocuenca@telefonica.com'
        sheet['G18'] = '0999586043'
    elif n2Responsable == 'Cristhian Vallejo':
        sheet['H18'] = 'cristhian.vallejovillalva@telefonica.com'
        sheet['G18'] = '0995386632'
    elif n2Responsable == 'Elizabeth Peñaloza':
        sheet['H18'] = 'julia.penaloza@telefonica.com'
        sheet['G18'] = '0998124791'
    elif n2Responsable == 'Fabricio Oleas':
        sheet['H18'] = 'luis.oleas@telefonica.com'
        sheet['G18'] = '0987537565'
    elif n2Responsable == 'Luis Chumo':
        sheet['H18'] = 'luis.chumo@telefonica.com'
        sheet['G18'] = '0999281697'
    elif n2Responsable ==  'Washington Chuqui':
        sheet['H18'] = 'washington.chuqui@telefonica.com'
        sheet['G18'] = '0995652744'
    elif n2Responsable ==  'Patricio Jordán':
        sheet['H18'] = 'patricio.jordanalvarez@telefonica.com'
        sheet['G18'] = '0989613952'
    elif n2Responsable ==  'Mauro Caluña':
        sheet['H18'] = 'mauro.calunatisalema@telefonica.com'
        sheet['G18'] = '0985796002'
    elif n2Responsable ==  'Xavier Ortiz':
        sheet['H18'] = 'xavier.ortiz@telefonica.com'
        sheet['G18'] = '0998542986'

    menu = ['MOP DIA', 'MOP NOCHE']
    choice = st.selectbox('Menu', menu)

    if choice == 'MOP NOCHE':
        st.title('MOP NOCHE')
        st.subheader('Actividades')
        with st.expander('1'):
            act1 = st.text_input('Actividad 1', value='N1 se contacta con el NOC para abrir la ventana de mantenimiento')
            sheet['B23'] = act1
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct1 = st.text_input('Hora I - 1', value='22:00')
                sheet['D23'] = horaInicioAct1
            with col2:
                horaFinAct1 = st.text_input('Hora F - 1', value='22:30')
                sheet['E23'] = horaFinAct1
            with col3:
                responsableAct1 = st.text_input('Responsable 1', value='N1/NOC')
                sheet['F23'] = responsableAct1
            with col4:
                afectacionAct1 = st.selectbox('Afectacion 1', ['N','Y'] )
                sheet['G23'] = afectacionAct1

        with st.expander('2'):
            act2 = st.text_input('Actividad 2', value='N1 se contacta con N2 para el inicio de la ventana')
            sheet['B24'] = act2
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct2 = st.text_input('Hora I - 2', value='22:30')
                sheet['D24'] = horaInicioAct2
            with col2:
                horaFinAct2 = st.text_input('Hora F - 2', value='23:00')
                sheet['E24'] = horaFinAct2
            with col3:
                responsableAct2 = st.text_input('Responsable 2', value='N1/N2')
                sheet['F24'] = responsableAct2
            with col4:
                afectacionAct2 = st.selectbox('Afectacion 2', ['N','Y'] )
                sheet['G24'] = afectacionAct2

        st.subheader('Inicio de Actividades')
        with st.expander('3'):
            act3 = st.text_input('Actividad 3', value='N2 inicia con las actividades de acuerdo a su Timeline')
            sheet['B26'] = act3
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct3 = st.text_input('Hora I - 3', value='00:00')
                sheet['D26'] = horaInicioAct3
            with col2:
                horaFinAct3 = st.text_input('Hora F - 3', value='03:00')
                sheet['E26'] = horaFinAct3
            with col3:
                responsableAct3 = st.text_input('Responsable 3', value='N2')
                sheet['F26'] = responsableAct3
            with col4:
                afectacionAct3 = st.selectbox('Afectacion 3', ['N','Y'] )
                sheet['G26'] = afectacionAct3

        with st.expander('4'):
            act4 = st.text_input('Actividad 4', value='N1 solicita avances a N2')
            sheet['B27'] = act4
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct4 = st.text_input('Hora I - 4', value='03:00')
                sheet['D27'] = horaInicioAct4
            with col2:
                horaFinAct4 = st.text_input('Hora F - 4', value='03:30')
                sheet['E27'] = horaFinAct4
            with col3:
                responsableAct4 = st.text_input('Responsable 4', value='N1/N2')
                sheet['F27'] = responsableAct4
            with col4:
                afectacionAct4 = st.selectbox('Afectacion 4', ['N','Y'] )
                sheet['G27'] = afectacionAct4
        
        with st.expander('5'):
            act5 = st.text_input('Actividad 5', value='N2 finaliza el mantenimiento')
            sheet['B28'] = act5
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct5 = st.text_input('Hora I - 5', value='03:30')
                sheet['D28'] = horaInicioAct5
            with col2:
                horaFinAct5 = st.text_input('Hora F - 5', value='04:00')
                sheet['E28'] = horaFinAct5
            with col3:
                responsableAct5 = st.text_input('Responsable 5', value='N2')
                sheet['F28'] = responsableAct5
            with col4:
                afectacionAct5 = st.selectbox('Afectacion 5', ['N','Y'] )
                sheet['G28'] = afectacionAct5
        
        with st.expander('6'):
            act6 = st.text_input('Actividad 6', value='N1 valida alarmas con el NOC y se cierra la ventana')
            sheet['B29'] = act6
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct6 = st.text_input('Hora I - 6', value='04:00')
                sheet['D29'] = horaInicioAct6
            with col2:
                horaFinAct6 = st.text_input('Hora F - 6', value='05:00')
                sheet['E29'] = horaFinAct6
            with col3:
                responsableAct6 = st.text_input('Responsable 6', value='N1/NOC')
                sheet['F29'] = responsableAct6
            with col4:
                afectacionAct6 = st.selectbox('Afectacion 6', ['N','Y'] )
                sheet['G29'] = afectacionAct6

        st.subheader('Rollback')
        rollback = st.selectbox('Rollback', ['NA', 'Inicio de ROLLBACK'])
        if rollback == 'Inicio de ROLLBACK':
            with st.expander('Rollback'):
                rollbackAct1 = st.text_input('Inicio de Rollback', value='Inicio de Rolback')
                sheet['B31'] = rollbackAct1
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback = st.text_input('Hora I - RLB', value='04:00')
                    sheet['D31'] = horaInicioRollback
                with col2:
                    horaFinRollback = st.text_input('Hora F - RLB', value='04:30')
                    sheet['E31'] = horaFinRollback
                with col3:
                    responsableRollback = st.text_input('Responsable RLB', value='N1/N2')
                    sheet['F31'] = responsableRollback
                with col4:
                    afectacionRollback = st.selectbox('Afectacion RLB', ['N','Y'] )
                    sheet['G31'] = afectacionRollback
                
                rollbackAct2 = st.text_input('Fin Rollback', value='N1 valida alarmas con el NOC y se cierra la ventana')
                sheet['B32'] = rollbackAct2
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback2 = st.text_input('Hora I - RLB2', value='04:30')
                    sheet['D32'] = horaInicioRollback2
                with col2:
                    horaFinRollback2 = st.text_input('Hora F - RLB2', value='05:00')
                    sheet['E32'] = horaFinRollback2
                with col3:
                    responsableRollback2 = st.text_input('Responsable RLB2', value='N1/NOC')
                    sheet['F32'] = responsableRollback2
                with col4:
                    afectacionRollback2 = st.selectbox('Afectacion RLB2', ['N','Y'] )
                    sheet['G32'] = afectacionRollback2

        st.subheader('Path de la documentacion')
        pathFolder = "\\\\10.112.47.36\\ingenieria\\o&m\\PLATAFORMAS Y REDES\\N1 O&M Redes y Plataformas\\"
        pathDocu = st.text_area('Path de la Documentacion', value=pathFolder)
        sheet['C34'] = pathDocu

        
        #if st.button('Guardar'):
        workbook.save('MOP-CRQ--NAME.xlsx')

        st.subheader('Descargar Archivo')
        if st.checkbox('Descargar:'):
            with open('MOP-CRQ--NAME.xlsx','rb') as fp:
                btn = st.download_button(label='Descargar Archivo', data=fp, file_name='MOP-CRQ--NAME.xlsx')

        st.write('All rights reserved. Developed by David Minango')    

    elif choice == 'MOP DIA':
        st.title('MOP DIA')

        st.subheader('Actividades')
        with st.expander('1'):
            act1 = st.text_input('Actividad 1', value='N1 se contacta con el NOC para abrir la ventana de mantenimiento')
            sheet['B23'] = act1
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct1 = st.text_input('Hora I - 1', value='09:00')
                sheet['D23'] = horaInicioAct1
            with col2:
                horaFinAct1 = st.text_input('Hora F - 1', value='09:30')
                sheet['E23'] = horaFinAct1
            with col3:
                responsableAct1 = st.text_input('Responsable 1', value='N1/NOC')
                sheet['F23'] = responsableAct1
            with col4:
                afectacionAct1 = st.selectbox('Afectacion 1', ['N','Y'] )
                sheet['G23'] = afectacionAct1

        with st.expander('2'):
            act2 = st.text_input('Actividad 2', value='N1 se contacta con N2 para el inicio de la ventana')
            sheet['B24'] = act2
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct2 = st.text_input('Hora I - 2', value='09:30')
                sheet['D24'] = horaInicioAct2
            with col2:
                horaFinAct2 = st.text_input('Hora F - 2', value='09:45')
                sheet['E24'] = horaFinAct2
            with col3:
                responsableAct2 = st.text_input('Responsable 2', value='N1/N2')
                sheet['F24'] = responsableAct2
            with col4:
                afectacionAct2 = st.selectbox('Afectacion 2', ['N','Y'] )
                sheet['G24'] = afectacionAct2

        st.subheader('Inicio de Actividades')
        with st.expander('3'):
            act3 = st.text_input('Actividad 3', value='N2 inicia con las actividades de acuerdo a su Timeline')
            sheet['B26'] = act3
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct3 = st.text_input('Hora I - 3', value='09:45')
                sheet['D26'] = horaInicioAct3
            with col2:
                horaFinAct3 = st.text_input('Hora F - 3', value='14:00')
                sheet['E26'] = horaFinAct3
            with col3:
                responsableAct3 = st.text_input('Responsable 3', value='N2')
                sheet['F26'] = responsableAct3
            with col4:
                afectacionAct3 = st.selectbox('Afectacion 3', ['N','Y'] )
                sheet['G26'] = afectacionAct3

        with st.expander('4'):
            act4 = st.text_input('Actividad 4', value='N1 solicita avances a N2')
            sheet['B27'] = act4
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct4 = st.text_input('Hora I - 4', value='14:00')
                sheet['D27'] = horaInicioAct4
            with col2:
                horaFinAct4 = st.text_input('Hora F - 4', value='15:00')
                sheet['E27'] = horaFinAct4
            with col3:
                responsableAct4 = st.text_input('Responsable 4', value='N1/N2')
                sheet['F27'] = responsableAct4
            with col4:
                afectacionAct4 = st.selectbox('Afectacion 4', ['N','Y'] )
                sheet['G27'] = afectacionAct4
        
        with st.expander('5'):
            act5 = st.text_input('Actividad 5', value='N2 finaliza el mantenimiento')
            sheet['B28'] = act5
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct5 = st.text_input('Hora I - 5', value='15:00')
                sheet['D28'] = horaInicioAct5
            with col2:
                horaFinAct5 = st.text_input('Hora F - 5', value='17:00')
                sheet['E28'] = horaFinAct5
            with col3:
                responsableAct5 = st.text_input('Responsable 5', value='N2')
                sheet['F28'] = responsableAct5
            with col4:
                afectacionAct5 = st.selectbox('Afectacion 5', ['N','Y'] )
                sheet['G28'] = afectacionAct5
        
        with st.expander('6'):
            act6 = st.text_input('Actividad 6', value='N1 valida alarmas con el NOC y se cierra la ventana')
            sheet['B29'] = act6
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct6 = st.text_input('Hora I - 6', value='17:00')
                sheet['D29'] = horaInicioAct6
            with col2:
                horaFinAct6 = st.text_input('Hora F - 6', value='18:00')
                sheet['E29'] = horaFinAct6
            with col3:
                responsableAct6 = st.text_input('Responsable 6', value='N1/NOC')
                sheet['F29'] = responsableAct6
            with col4:
                afectacionAct6 = st.selectbox('Afectacion 6', ['N','Y'] )
                sheet['G29'] = afectacionAct6

        st.subheader('Rollback')
        rollback = st.selectbox('Rollback', ['NA', 'Inicio de ROLLBACK'])
        if rollback == 'Inicio de ROLLBACK':
            with st.expander('Rollback'):
                rollbackAct1 = st.text_input('Inicio de Rollback', value='Inicio de Rolback')
                sheet['B31'] = rollbackAct1
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback = st.text_input('Hora I - RLB', value='17:00')
                    sheet['D31'] = horaInicioRollback
                with col2:
                    horaFinRollback = st.text_input('Hora F - RLB', value='17:30')
                    sheet['E31'] = horaFinRollback
                with col3:
                    responsableRollback = st.text_input('Responsable RLB', value='N1/N2')
                    sheet['F31'] = responsableRollback
                with col4:
                    afectacionRollback = st.selectbox('Afectacion RLB', ['N','Y'] )
                    sheet['G31'] = afectacionRollback
                
                rollbackAct2 = st.text_input('Fin Rollback', value='N1 valida alarmas con el NOC y se cierra la ventana')
                sheet['B32'] = rollbackAct2
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback2 = st.text_input('Hora I - RLB2', value='17:30')
                    sheet['D32'] = horaInicioRollback2
                with col2:
                    horaFinRollback2 = st.text_input('Hora F - RLB2', value='18:00')
                    sheet['E32'] = horaFinRollback2
                with col3:
                    responsableRollback2 = st.text_input('Responsable RLB2', value='N1/NOC')
                    sheet['F32'] = responsableRollback2
                with col4:
                    afectacionRollback2 = st.selectbox('Afectacion RLB2', ['N','Y'] )
                    sheet['G32'] = afectacionRollback2


        st.subheader('Path de la documentacion')
        pathFolder = "\\\\10.112.47.36\\ingenieria\\o&m\\PLATAFORMAS Y REDES\\N1 O&M Redes y Plataformas\\"
        pathDocu = st.text_area('Path de la Documentacion', value=pathFolder)
        sheet['C34'] = pathDocu

        
        #if st.button('Guardar'):
        workbook.save('MOP-CRQ--NAME.xlsx')

        st.subheader('Descargar Archivo')
        if st.checkbox('Descargar:'):
            with open('MOP-CRQ--NAME.xlsx','rb') as fp:
                btn = st.download_button(label='Descargar Archivo', data=fp, file_name='MOP-CRQ--NAME.xlsx')
################################################################################################################
################################################################################################################
def mopProveedor():

    workbook = openpyxl.load_workbook('MOP-CRQ_Prove.xlsx')
    sheets = workbook.sheetnames
    sheet = workbook.active

    nombreCRQ = st.text_input('Nombre del Mantenimiento:')
    sheet['B9'] = nombreCRQ
    sheet['C11'] = nombreCRQ
    CRQ = nombreCRQ[4:10]
    numCRQ = st.text_input('Numero de CRQ:', value=CRQ)
    if numCRQ.isdigit() == True:
        sheet['D10'] = numCRQ
    else:
        st.warning('Ingrese un numero valido')
    
    impacto = st.text_input('Impacto / Alcance', value='Ninguno')
    sheet['C13'] = impacto

    n2Responsable = st.selectbox('N2', ['Carlos Castillo','Cristhian Vallejo', 'Elizabeth Peñaloza', 'Fabricio Oleas','Luis Chumo', 'Washington Chuqui', 'Patricio Jordán', 'Mauro Caluña', 'Xavier Ortiz'])
    sheet['E18'] = n2Responsable
    if n2Responsable == 'Carlos Castillo':
        sheet['H18'] = 'carlos.castillocuenca@telefonica.com'
        sheet['G18'] = '0999586043'
    elif n2Responsable == 'Cristhian Vallejo':
        sheet['H18'] = 'cristhian.vallejovillalva@telefonica.com'
        sheet['G18'] = '0995386632'
    elif n2Responsable == 'Elizabeth Peñaloza':
        sheet['H18'] = 'julia.penaloza@telefonica.com'
        sheet['G18'] = '0998124791'
    elif n2Responsable == 'Fabricio Oleas':
        sheet['H18'] = 'luis.oleas@telefonica.com'
        sheet['G18'] = '0987537565'
    elif n2Responsable == 'Luis Chumo':
        sheet['H18'] = 'luis.chumo@telefonica.com'
        sheet['G18'] = '0999281697'
    elif n2Responsable ==  'Washington Chuqui':
        sheet['H18'] = 'washington.chuqui@telefonica.com'
        sheet['G18'] = '0995652744'
    elif n2Responsable ==  'Patricio Jordán':
        sheet['H18'] = 'patricio.jordanalvarez@telefonica.com'
        sheet['G18'] = '0989613952'
    elif n2Responsable ==  'Mauro Caluña':
        sheet['H18'] = 'mauro.calunatisalema@telefonica.com'
        sheet['G18'] = '0985796002'
    elif n2Responsable ==  'Xavier Ortiz':
        sheet['H18'] = 'xavier.ortiz@telefonica.com'
        sheet['G18'] = '0998542986'

    proveedor = st.text_input('Proveedor')
    sheet['E19'] = proveedor
    proveedorNum = st.text_input('Proveedor Telefono')
    sheet['G19'] = proveedorNum
    proveedorEmail = st.text_input('Proveedor Correo')
    sheet['H19'] = proveedorEmail
    
    menu = ['MOP DIA', 'MOP NOCHE']
    choice = st.selectbox('Menu', menu)

    if choice == 'MOP NOCHE':
        st.title('MOP NOCHE')
        st.subheader('Actividades')
        with st.expander('1'):
            act1 = st.text_input('Actividad 1', value='N1 se contacta con el NOC para abrir la ventana de mantenimiento')
            sheet['B23'] = act1
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct1 = st.text_input('Hora I - 1', value='22:00')
                sheet['D23'] = horaInicioAct1
            with col2:
                horaFinAct1 = st.text_input('Hora F - 1', value='22:30')
                sheet['E23'] = horaFinAct1
            with col3:
                responsableAct1 = st.text_input('Responsable 1', value='N1/NOC')
                sheet['F23'] = responsableAct1
            with col4:
                afectacionAct1 = st.selectbox('Afectacion 1', ['N','Y'] )
                sheet['G23'] = afectacionAct1

        with st.expander('2'):
            act2 = st.text_input('Actividad 2', value='N1 se contacta con Proveedor para el inicio de la ventana')
            sheet['B24'] = act2
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct2 = st.text_input('Hora I - 2', value='22:30')
                sheet['D24'] = horaInicioAct2
            with col2:
                horaFinAct2 = st.text_input('Hora F - 2', value='23:00')
                sheet['E24'] = horaFinAct2
            with col3:
                responsableAct2 = st.text_input('Responsable 2', value='N1/Proveedor')
                sheet['F24'] = responsableAct2
            with col4:
                afectacionAct2 = st.selectbox('Afectacion 2', ['N','Y'] )
                sheet['G24'] = afectacionAct2

        st.subheader('Inicio de Actividades')
        with st.expander('3'):
            act3 = st.text_input('Actividad 3', value='Proveedor inicia con las actividades de acuerdo a su Timeline')
            sheet['B26'] = act3
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct3 = st.text_input('Hora I - 3', value='00:00')
                sheet['D26'] = horaInicioAct3
            with col2:
                horaFinAct3 = st.text_input('Hora F - 3', value='03:00')
                sheet['E26'] = horaFinAct3
            with col3:
                responsableAct3 = st.text_input('Responsable 3', value='Proveedor')
                sheet['F26'] = responsableAct3
            with col4:
                afectacionAct3 = st.selectbox('Afectacion 3', ['N','Y'] )
                sheet['G26'] = afectacionAct3

        with st.expander('4'):
            act4 = st.text_input('Actividad 4', value='N1 solicita avances a Proveedor')
            sheet['B27'] = act4
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct4 = st.text_input('Hora I - 4', value='03:00')
                sheet['D27'] = horaInicioAct4
            with col2:
                horaFinAct4 = st.text_input('Hora F - 4', value='03:30')
                sheet['E27'] = horaFinAct4
            with col3:
                responsableAct4 = st.text_input('Responsable 4', value='N1/Proveedor')
                sheet['F27'] = responsableAct4
            with col4:
                afectacionAct4 = st.selectbox('Afectacion 4', ['N','Y'] )
                sheet['G27'] = afectacionAct4
        
        with st.expander('5'):
            act5 = st.text_input('Actividad 5', value='Proveedor finaliza el mantenimiento')
            sheet['B28'] = act5
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct5 = st.text_input('Hora I - 5', value='03:30')
                sheet['D28'] = horaInicioAct5
            with col2:
                horaFinAct5 = st.text_input('Hora F - 5', value='04:00')
                sheet['E28'] = horaFinAct5
            with col3:
                responsableAct5 = st.text_input('Responsable 5', value='Proveedor')
                sheet['F28'] = responsableAct5
            with col4:
                afectacionAct5 = st.selectbox('Afectacion 5', ['N','Y'] )
                sheet['G28'] = afectacionAct5
        
        with st.expander('6'):
            act6 = st.text_input('Actividad 6', value='N1 valida alarmas con el NOC y se cierra la ventana')
            sheet['B29'] = act6
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct6 = st.text_input('Hora I - 6', value='04:00')
                sheet['D29'] = horaInicioAct6
            with col2:
                horaFinAct6 = st.text_input('Hora F - 6', value='05:00')
                sheet['E29'] = horaFinAct6
            with col3:
                responsableAct6 = st.text_input('Responsable 6', value='N1/NOC')
                sheet['F29'] = responsableAct6
            with col4:
                afectacionAct6 = st.selectbox('Afectacion 6', ['N','Y'] )
                sheet['G29'] = afectacionAct6

        st.subheader('Rollback')
        rollback = st.selectbox('Rollback', ['NA', 'Inicio de ROLLBACK'])
        if rollback == 'Inicio de ROLLBACK':
            with st.expander('Rollback'):
                rollbackAct1 = st.text_input('Inicio de Rollback', value='Inicio de Rolback')
                sheet['B31'] = rollbackAct1
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback = st.text_input('Hora I - RLB', value='04:00')
                    sheet['D31'] = horaInicioRollback
                with col2:
                    horaFinRollback = st.text_input('Hora F - RLB', value='04:30')
                    sheet['E31'] = horaFinRollback
                with col3:
                    responsableRollback = st.text_input('Responsable RLB', value='N1/Proveedor')
                    sheet['F31'] = responsableRollback
                with col4:
                    afectacionRollback = st.selectbox('Afectacion RLB', ['N','Y'] )
                    sheet['G31'] = afectacionRollback
                
                rollbackAct2 = st.text_input('Fin Rollback', value='N1 valida alarmas con el NOC y se cierra la ventana')
                sheet['B32'] = rollbackAct2
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback2 = st.text_input('Hora I - RLB2', value='04:30')
                    sheet['D32'] = horaInicioRollback2
                with col2:
                    horaFinRollback2 = st.text_input('Hora F - RLB2', value='05:00')
                    sheet['E32'] = horaFinRollback2
                with col3:
                    responsableRollback2 = st.text_input('Responsable RLB2', value='N1/NOC')
                    sheet['F32'] = responsableRollback2
                with col4:
                    afectacionRollback2 = st.selectbox('Afectacion RLB2', ['N','Y'] )
                    sheet['G32'] = afectacionRollback2

        st.subheader('Path de la documentacion')
        pathFolder = "\\\\10.112.47.36\\ingenieria\\o&m\\PLATAFORMAS Y REDES\\N1 O&M Redes y Plataformas\\"
        pathDocu = st.text_area('Path de la Documentacion', value=pathFolder)
        sheet['C34'] = pathDocu

        
        #if st.button('Guardar'):
        workbook.save('MOP-CRQ--NAME2.xlsx')

        st.subheader('Descargar Archivo')
        if st.checkbox('Descargar:'):
            with open('MOP-CRQ--NAME2.xlsx','rb') as fp:
                btn = st.download_button(label='Descargar Archivo', data=fp, file_name='MOP-CRQ--NAME.xlsx')

        st.write('All rights reserved. Developed by David Minango')

    elif choice == 'MOP DIA':
        st.title('MOP DIA')

        st.subheader('Actividades')
        with st.expander('1'):
            act1 = st.text_input('Actividad 1', value='N1 se contacta con el NOC para abrir la ventana de mantenimiento')
            sheet['B23'] = act1
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct1 = st.text_input('Hora I - 1', value='09:00')
                sheet['D23'] = horaInicioAct1
            with col2:
                horaFinAct1 = st.text_input('Hora F - 1', value='09:30')
                sheet['E23'] = horaFinAct1
            with col3:
                responsableAct1 = st.text_input('Responsable 1', value='N1/NOC')
                sheet['F23'] = responsableAct1
            with col4:
                afectacionAct1 = st.selectbox('Afectacion 1', ['N','Y'] )
                sheet['G23'] = afectacionAct1

        with st.expander('2'):
            act2 = st.text_input('Actividad 2', value='N1 se contacta con Proveedor para el inicio de la ventana')
            sheet['B24'] = act2
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct2 = st.text_input('Hora I - 2', value='09:30')
                sheet['D24'] = horaInicioAct2
            with col2:
                horaFinAct2 = st.text_input('Hora F - 2', value='09:45')
                sheet['E24'] = horaFinAct2
            with col3:
                responsableAct2 = st.text_input('Responsable 2', value='N1/Proveedor')
                sheet['F24'] = responsableAct2
            with col4:
                afectacionAct2 = st.selectbox('Afectacion 2', ['N','Y'] )
                sheet['G24'] = afectacionAct2

        st.subheader('Inicio de Actividades')
        with st.expander('3'):
            act3 = st.text_input('Actividad 3', value='Proveedor inicia con las actividades de acuerdo a su Timeline')
            sheet['B26'] = act3
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct3 = st.text_input('Hora I - 3', value='09:45')
                sheet['D26'] = horaInicioAct3
            with col2:
                horaFinAct3 = st.text_input('Hora F - 3', value='14:00')
                sheet['E26'] = horaFinAct3
            with col3:
                responsableAct3 = st.text_input('Responsable 3', value='Proveedor')
                sheet['F26'] = responsableAct3
            with col4:
                afectacionAct3 = st.selectbox('Afectacion 3', ['N','Y'] )
                sheet['G26'] = afectacionAct3

        with st.expander('4'):
            act4 = st.text_input('Actividad 4', value='N1 solicita avances a Proveedor')
            sheet['B27'] = act4
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct4 = st.text_input('Hora I - 4', value='14:00')
                sheet['D27'] = horaInicioAct4
            with col2:
                horaFinAct4 = st.text_input('Hora F - 4', value='15:00')
                sheet['E27'] = horaFinAct4
            with col3:
                responsableAct4 = st.text_input('Responsable 4', value='N1/Proveedor')
                sheet['F27'] = responsableAct4
            with col4:
                afectacionAct4 = st.selectbox('Afectacion 4', ['N','Y'] )
                sheet['G27'] = afectacionAct4
        
        with st.expander('5'):
            act5 = st.text_input('Actividad 5', value='Proveedor finaliza el mantenimiento')
            sheet['B28'] = act5
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct5 = st.text_input('Hora I - 5', value='15:00')
                sheet['D28'] = horaInicioAct5
            with col2:
                horaFinAct5 = st.text_input('Hora F - 5', value='17:00')
                sheet['E28'] = horaFinAct5
            with col3:
                responsableAct5 = st.text_input('Responsable 5', value='Proveedor')
                sheet['F28'] = responsableAct5
            with col4:
                afectacionAct5 = st.selectbox('Afectacion 5', ['N','Y'] )
                sheet['G28'] = afectacionAct5
        
        with st.expander('6'):
            act6 = st.text_input('Actividad 6', value='N1 valida alarmas con el NOC y se cierra la ventana')
            sheet['B29'] = act6
            col1,col2,col3,col4 = st.columns(4)       
            with col1:
                horaInicioAct6 = st.text_input('Hora I - 6', value='17:00')
                sheet['D29'] = horaInicioAct6
            with col2:
                horaFinAct6 = st.text_input('Hora F - 6', value='18:00')
                sheet['E29'] = horaFinAct6
            with col3:
                responsableAct6 = st.text_input('Responsable 6', value='N1/NOC')
                sheet['F29'] = responsableAct6
            with col4:
                afectacionAct6 = st.selectbox('Afectacion 6', ['N','Y'] )
                sheet['G29'] = afectacionAct6

        st.subheader('Rollback')
        rollback = st.selectbox('Rollback', ['NA', 'Inicio de ROLLBACK'])
        if rollback == 'Inicio de ROLLBACK':
            with st.expander('Rollback'):
                rollbackAct1 = st.text_input('Inicio de Rollback', value='Inicio de Rolback')
                sheet['B31'] = rollbackAct1
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback = st.text_input('Hora I - RLB', value='17:00')
                    sheet['D31'] = horaInicioRollback
                with col2:
                    horaFinRollback = st.text_input('Hora F - RLB', value='17:30')
                    sheet['E31'] = horaFinRollback
                with col3:
                    responsableRollback = st.text_input('Responsable RLB', value='N1/Proveedor')
                    sheet['F31'] = responsableRollback
                with col4:
                    afectacionRollback = st.selectbox('Afectacion RLB', ['N','Y'] )
                    sheet['G31'] = afectacionRollback
                
                rollbackAct2 = st.text_input('Fin Rollback', value='N1 valida alarmas con el NOC y se cierra la ventana')
                sheet['B32'] = rollbackAct2
                col1,col2,col3,col4 = st.columns(4)       
                with col1:
                    horaInicioRollback2 = st.text_input('Hora I - RLB2', value='17:30')
                    sheet['D32'] = horaInicioRollback2
                with col2:
                    horaFinRollback2 = st.text_input('Hora F - RLB2', value='18:00')
                    sheet['E32'] = horaFinRollback2
                with col3:
                    responsableRollback2 = st.text_input('Responsable RLB2', value='N1/NOC')
                    sheet['F32'] = responsableRollback2
                with col4:
                    afectacionRollback2 = st.selectbox('Afectacion RLB2', ['N','Y'] )
                    sheet['G32'] = afectacionRollback2


        st.subheader('Path de la documentacion')
        pathFolder = "\\\\10.112.47.36\\ingenieria\\o&m\\PLATAFORMAS Y REDES\\N1 O&M Redes y Plataformas\\"
        pathDocu = st.text_area('Path de la Documentacion', value=pathFolder)
        sheet['C34'] = pathDocu

        
        #if st.button('Guardar'):
        workbook.save('MOP-CRQ--NAME2.xlsx')

        st.subheader('Descargar Archivo')
        if st.checkbox('Descargar:'):
            with open('MOP-CRQ--NAME2.xlsx','rb') as fp:
                btn = st.download_button(label='Descargar Archivo', data=fp, file_name='MOP-CRQ--NAME.xlsx')

################################################################################################################
################################################################################################################
def mopPreaprobadas():

    workbook = openpyxl.load_workbook('MopsPreaprobadas.xlsx')
    sheets = workbook.sheetnames
    sheet = workbook.active

    objetivo = st.text_input('Objetivo:')
    sheet['C7'] = objetivo

    nomMantenimiento = st.text_input('Nombre de mantenimiento:')
    sheet['C8'] = nomMantenimiento

    descripcion = st.text_input('Descripción de impacto:', value='No se Espera Afectación')
    sheet['C9'] = descripcion

    col1, col2 = st.columns(2)
    with col1:
        fechaIniciomtto = st.date_input('Fecha Inicio MTTO')
        sheet['F13'] = fechaIniciomtto.strftime('%d-%b-%Y')
        fechaFinmtto = st.date_input('Fecha Fin MTTO')
        sheet['F14'] = fechaFinmtto.strftime('%d-%b-%Y')
        fechaInicioAfectacion = st.date_input('Fecha Inicio Afectación')
        sheet['F15'] = fechaInicioAfectacion.strftime('%d-%b-%Y')
        fechaFinAfectacion = st.date_input('Fecha Fin Afectación')
        sheet['F16'] = fechaFinAfectacion.strftime('%d-%b-%Y')
        fechaInicioRollback = st.date_input('Fecha Inicio Rollback')
        sheet['F17'] = fechaInicioRollback.strftime('%d-%b-%Y')
        fechaFinRollback = st.date_input('Fecha Fin Rollback')
        sheet['F18'] = fechaFinRollback.strftime('%d-%b-%Y')

    with col2:
        horaIniciomtto = st.text_input('Hora Inicio MTTO', value='00:00')
        sheet['G13'] = horaIniciomtto
        horaFinmtto = st.text_input('Hora Fin MTTO', value='05:00')
        sheet['G14'] = horaFinmtto
        horaInicioAfectacion = st.text_input('Hora Inicio Afectacion', value='00:00')
        sheet['G15'] = horaInicioAfectacion
        horaFinAfectacion = st.text_input('Hora Fin Afectacion',value='00:00')
        sheet['G16'] = horaFinAfectacion
        horaInicioRollback = st.text_input('Hora Inicio Rollback', value='00:00')
        sheet['G17'] = horaInicioRollback
        horaFinRollback = st.text_input('Hora Fin Rollback',value='00:00')
        sheet['G18'] = horaFinRollback

        sheet['C13'] = str(fechaIniciomtto.strftime('%d-%m-%Y')) +" "+ str(horaIniciomtto)
        sheet['C14'] = str(fechaFinmtto.strftime('%d-%m-%Y')) +" "+ str(horaFinmtto)
        sheet['C15'] = str(fechaInicioAfectacion.strftime('%d-%m-%Y')) +" "+ str(horaInicioAfectacion)
        sheet['C16'] = str(fechaFinAfectacion.strftime('%d-%m-%Y')) +" "+ str(horaFinAfectacion)
        sheet['C17'] = str(fechaInicioRollback.strftime('%d-%m-%Y')) +" "+ str(horaInicioRollback)
        sheet['C18'] = str(fechaFinRollback.strftime('%d-%m-%Y')) +" "+ str(horaFinRollback)

    ppmProyecto = st.text_input('PPM (Proyecto)',value='EX-')
    sheet['C19'] = ppmProyecto
    col12,col22 = st.columns(2)
    with col12:
        resOtecel = st.text_input('Responsable Otecel', value='N2 O&M')
        sheet['C20'] = resOtecel
        resProveedor = st.text_input('Responsable Proveedor', value='N1 PAP')
        sheet['C23'] = resProveedor
        

    with col22:
        tlfOtecel = st.text_input('Telefono Otecel', value='0999729993')
        if tlfOtecel.isdigit() == True:
            sheet['C21'] = tlfOtecel
        else:
            st.warning('Ingrese un numero valido')

        tlfProveedor = st.text_input('Telefono Proveedor', value='0999993070')
        if tlfProveedor.isdigit() == True:
            sheet['C24'] = tlfProveedor
        else:
            st.warning('Ingrese un numero valido')
        
        emailProveedor = st.text_input('Email Proveedor', value='n1pap.ec@telefonica.com')
        sheet['C25'] = emailProveedor

    if st.button('Guardar'):
        workbook.save('MOP_Preaprobadas.xlsx')

    st.subheader('Descargar Archivo')
    if st.checkbox('Descargar:'):
        with open('MOP_Preaprobadas.xlsx','rb') as fp:
            btn = st.download_button(label='Descargar Archivo', data=fp, file_name='MOP_Preaprobadas.xlsx')
################################################################################################################

def mainMops():

    #st.image('proconty.png', use_column_width=True)
    st.title('Pasos a Producción')
    
    #st.sidebar.image('proconty.png', width=300)
    
    menu = ['MOP Gestión N2 - Generico','MOP Proveedor - Generico', 'MOP Preaprobadas']
    choice = st.sidebar.selectbox('Menu - MOPs', menu)

    if choice == 'MOP Gestión N2 - Generico':
        st.subheader('MOP N2')
        mopMantenimientos()
    
    elif choice == 'MOP Proveedor - Generico':
        st.subheader('MOP Proveedores')
        mopProveedor()

    elif choice == 'MOP Preaprobadas':
        st.subheader('MOP Preaprobadas')
        mopPreaprobadas()


    #st.sidebar.markdown('[MENU PRINCIPAL](https://proconty-pap.herokuapp.com/)')
    
    
